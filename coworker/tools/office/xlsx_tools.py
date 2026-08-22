"""Excel workbooks — read, write, and edit .xlsx files.

Two failure modes drive this design.

*Context flooding*: a real workbook has more rows than a context window. ``read_workbook``
therefore windows rows exactly as ``read_file`` windows lines, reports ``total_rows``, and
tells the model how to continue. It never returns a whole sheet by default.

*Silent formula destruction*: rewriting a sheet to change one cell discards every formula,
chart, and format in it. ``edit_workbook`` writes individual cells through openpyxl and
saves, so everything it did not touch survives. ``read_workbook(formulas=True)`` shows the
formulas themselves, so the model can see what it would be standing on.
"""

from __future__ import annotations

from typing import Any

from ... import deliverable_check
from ._common import clip, decorate, guard, require
from .paths import context_roots, display_path, resolve_read, resolve_write

_DEFAULT_ROWS = 100
_MAX_ROWS = 1000

_READ_SCHEMA = {
    "type": "function",
    "function": {
        "name": "read_workbook",
        "description": (
            "Read an Excel (.xlsx) sheet as rows. Returns the sheet names, the header row, and "
            "a window of data rows with total_rows so you know how much is left — pass "
            "start_row to continue. Set formulas=true to see cell formulas instead of their "
            "cached values. For statistics or aggregation over a large sheet, prefer "
            "inspect_data or run_python (pandas) instead of reading every row. Read-only."
        ),
        "parameters": {
            "type": "object",
            "properties": {
                "path": {"type": "string", "description": "The .xlsx file to read."},
                "sheet": {
                    "type": "string",
                    "description": "Sheet name (default: the first sheet).",
                },
                "start_row": {
                    "type": "integer",
                    "description": "First data row to return, 1-based (default 1).",
                },
                "max_rows": {
                    "type": "integer",
                    "description": f"How many rows (default {_DEFAULT_ROWS}, max {_MAX_ROWS}).",
                },
                "formulas": {
                    "type": "boolean",
                    "description": "Return formulas rather than cached values (default false).",
                },
            },
            "required": ["path"],
        },
    },
}

_WRITE_SCHEMA = {
    "type": "function",
    "function": {
        "name": "write_workbook",
        "description": (
            "Create or overwrite an Excel (.xlsx) workbook from rows. Pass one sheet, or "
            "several via 'sheets'. The first row is styled as a header. Use this for any "
            "spreadsheet deliverable — do NOT write a script to do it."
        ),
        "parameters": {
            "type": "object",
            "properties": {
                "path": {"type": "string", "description": "Destination .xlsx path."},
                "rows": {
                    "type": "array",
                    "description": "Rows for a single sheet; each row is a list of cell values.",
                    "items": {"type": "array", "items": {}},
                },
                "sheet": {
                    "type": "string",
                    "description": "Sheet name when using 'rows' (default 'Sheet1').",
                },
                "sheets": {
                    "type": "array",
                    "description": "Multiple sheets, instead of 'rows'.",
                    "items": {
                        "type": "object",
                        "properties": {
                            "name": {"type": "string"},
                            "rows": {"type": "array", "items": {"type": "array", "items": {}}},
                        },
                        "required": ["name", "rows"],
                    },
                },
                "header": {
                    "type": "boolean",
                    "description": "Bold the first row and freeze it (default true).",
                },
            },
            "required": ["path"],
        },
    },
}

_EDIT_SCHEMA = {
    "type": "function",
    "function": {
        "name": "edit_workbook",
        "description": (
            "Set individual cells in an existing Excel workbook, in place. Everything not "
            "written — formulas, formats, charts, other sheets — is preserved. A value "
            "starting with '=' is written as a formula."
        ),
        "parameters": {
            "type": "object",
            "properties": {
                "path": {"type": "string", "description": "The .xlsx file to edit."},
                "sheet": {
                    "type": "string",
                    "description": "Sheet name (default: the first sheet).",
                },
                "cells": {
                    "type": "array",
                    "description": "Cell writes to apply.",
                    "items": {
                        "type": "object",
                        "properties": {
                            "cell": {
                                "type": "string",
                                "description": "A1-style reference, e.g. 'B7'.",
                            },
                            "value": {"description": "New value; '=SUM(A1:A9)' writes a formula."},
                        },
                        "required": ["cell", "value"],
                    },
                },
            },
            "required": ["path", "cells"],
        },
    },
}


def _cell(value: Any) -> Any:
    """Bound one model-visible cell; dates and numbers keep their type."""
    if value is None:
        return None
    if isinstance(value, (int, float, bool)):
        return value
    return clip(str(value))


def xlsx_tools(context: Any) -> list:
    roots = context_roots(context)

    @guard
    def read_workbook(
        path: str,
        sheet: str = "",
        start_row: int = 1,
        max_rows: int = _DEFAULT_ROWS,
        formulas: bool = False,
    ) -> dict[str, Any]:
        openpyxl = require("openpyxl", "openpyxl")
        target = resolve_read(path, roots)
        if not target.is_file():
            raise FileNotFoundError(display_path(target, roots))

        # data_only=True gives the value Excel last cached; False gives the formula text.
        book = openpyxl.load_workbook(
            str(target), data_only=not formulas, read_only=True
        )
        try:
            names = list(book.sheetnames)
            if sheet:
                if sheet not in names:
                    raise KeyError(f"no sheet named {sheet!r}; available: {', '.join(names)}")
                ws = book[sheet]
            else:
                ws = book[names[0]]

            begin = start_row if isinstance(start_row, int) and start_row > 0 else 1
            count = max_rows if isinstance(max_rows, int) and max_rows > 0 else _DEFAULT_ROWS
            count = min(count, _MAX_ROWS)

            rows: list[list[Any]] = []
            header: list[Any] = []
            total = 0
            for i, row in enumerate(ws.iter_rows(values_only=True), 1):
                total = i
                if i == 1:
                    header = [_cell(v) for v in row]
                    continue
                data_index = i - 1  # 1-based among data rows (header excluded)
                if data_index < begin or len(rows) >= count:
                    continue
                rows.append([_cell(v) for v in row])

            data_rows = max(total - 1, 0)
            result: dict[str, Any] = {
                "path": display_path(target, roots),
                "sheet": ws.title,
                "sheets": names,
                "header": header,
                "rows": rows,
                "total_rows": data_rows,
                "start_row": begin,
            }
            end = begin + len(rows)
            if end - 1 < data_rows:
                result["note"] = (
                    f"showing data rows {begin}-{end - 1} of {data_rows}; call again with "
                    f"start_row={end} to continue, or use inspect_data/run_python to "
                    "analyse the sheet without reading it all"
                )
            return result
        finally:
            book.close()

    @guard
    def write_workbook(
        path: str,
        rows: Any = None,
        sheet: str = "Sheet1",
        sheets: Any = None,
        header: bool = True,
    ) -> dict[str, Any]:
        openpyxl = require("openpyxl", "openpyxl")
        from openpyxl.styles import Font

        target = resolve_write(path, roots)
        if sheets is None and rows is None:
            raise ValueError("pass either 'rows' (one sheet) or 'sheets' (several)")

        plan = (
            list(sheets)
            if sheets is not None
            else [{"name": sheet or "Sheet1", "rows": rows or []}]
        )

        book = openpyxl.Workbook()
        book.remove(book.active)
        written = 0
        for entry in plan:
            if not isinstance(entry, dict):
                raise ValueError("each sheet must be an object with 'name' and 'rows'")
            name = str(entry.get("name") or "Sheet1")[:31]  # Excel's sheet-name limit
            data = entry.get("rows") or []
            if not isinstance(data, list):
                raise ValueError(f"sheet {name!r}: 'rows' must be a list of rows")
            ws = book.create_sheet(title=name)
            for row in data:
                ws.append(list(row) if isinstance(row, (list, tuple)) else [row])
                written += 1
            if header and data:
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                ws.freeze_panes = "A2"
            # Width from the widest value per column, so the deliverable opens readable
            # rather than showing ####. Bounded so one long cell can't blow the layout out.
            for column in ws.columns:
                longest = max((len(str(c.value)) for c in column if c.value is not None), default=0)
                ws.column_dimensions[column[0].column_letter].width = min(max(longest + 2, 8), 60)

        target.parent.mkdir(parents=True, exist_ok=True)
        book.save(str(target))
        return deliverable_check.attach(
            {
                "path": display_path(target, roots),
                "sheets": [str(e.get("name")) for e in plan],
                "rows_written": written,
                "bytes": target.stat().st_size,
            },
            target,
        )

    @guard
    def edit_workbook(path: str, cells: list, sheet: str = "") -> dict[str, Any]:
        openpyxl = require("openpyxl", "openpyxl")
        target = resolve_write(path, roots)
        if not target.is_file():
            raise FileNotFoundError(display_path(target, roots))
        if not isinstance(cells, list) or not cells:
            raise ValueError("'cells' must be a non-empty list")

        # data_only=False is essential: loading with data_only=True and saving would replace
        # every formula in the workbook with its last cached value.
        book = openpyxl.load_workbook(str(target), data_only=False)
        names = list(book.sheetnames)
        if sheet:
            if sheet not in names:
                raise KeyError(f"no sheet named {sheet!r}; available: {', '.join(names)}")
            ws = book[sheet]
        else:
            ws = book[names[0]]

        applied = 0
        for item in cells:
            if not isinstance(item, dict):
                raise ValueError("each cell write must be an object with 'cell' and 'value'")
            ref = str(item.get("cell") or "").strip().upper()
            if not ref:
                raise ValueError("each cell write needs a 'cell' reference like 'B7'")
            try:
                ws[ref] = item.get("value")
            except (ValueError, TypeError) as exc:
                raise ValueError(f"cannot write {ref}: {exc}") from exc
            applied += 1

        book.save(str(target))
        return {"path": display_path(target, roots), "sheet": ws.title, "cells_written": applied}

    return [
        decorate(read_workbook, name="read_workbook", schema=_READ_SCHEMA),
        decorate(
            write_workbook,
            name="write_workbook",
            schema=_WRITE_SCHEMA,
            risk="medium",
            capabilities=["write"],
        ),
        decorate(
            edit_workbook,
            name="edit_workbook",
            schema=_EDIT_SCHEMA,
            risk="medium",
            capabilities=["write"],
        ),
    ]
